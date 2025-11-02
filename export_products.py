import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime

# 数据库连接
DATABASE_URL = os.environ.get('DATABASE_URL', '')
if not DATABASE_URL:
    print("Error: DATABASE_URL not found")
    exit(1)

# 解析数据库URL
# mysql://user:password@host:port/database?ssl=...
from urllib.parse import urlparse

parsed = urlparse(DATABASE_URL)
user = parsed.username
password = parsed.password
host = parsed.hostname
port = parsed.port or 3306
database = parsed.path.lstrip('/')

# 移除SSL参数
if '?' in database:
    database = database.split('?')[0]

print(f"Connecting to database: {database}@{host}:{port}")

# 连接数据库
conn = mysql.connector.connect(
    host=host,
    port=int(port),
    user=user,
    password=password,
    database=database
)

cursor = conn.cursor(dictionary=True)

# 查询所有产品
query = """
SELECT 
  productId as 'ROWELL产品编号',
  partNumber as '原厂Part Number',
  brand as '品牌',
  name as '产品名称',
  description as '产品描述',
  specifications as '产品规格',
  productType as '产品类型',
  status as '状态'
FROM products 
ORDER BY brand, partNumber
"""

cursor.execute(query)
products = cursor.fetchall()

print(f"查询到 {len(products)} 个产品")

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ROWELL产品目录"

# 设置列标题
headers = ['品牌', 'ROWELL产品编号', '原厂Part Number', '产品名称', '产品描述', '产品规格', '产品类型', '状态']
ws.append(headers)

# 设置标题样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 添加数据
for product in products:
    row = [
        product.get('品牌', ''),
        product.get('ROWELL产品编号', ''),
        product.get('原厂Part Number', ''),
        product.get('产品名称', ''),
        product.get('产品描述', ''),
        product.get('产品规格', ''),
        product.get('产品类型', ''),
        product.get('状态', '')
    ]
    ws.append(row)

# 设置列宽
column_widths = {
    'A': 20,  # 品牌
    'B': 20,  # ROWELL产品编号
    'C': 20,  # 原厂Part Number
    'D': 40,  # 产品名称
    'E': 60,  # 产品描述
    'F': 40,  # 产品规格
    'G': 25,  # 产品类型
    'H': 12   # 状态
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 设置数据行样式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# 保存文件
output_file = '/home/ubuntu/ROWELL_产品目录_完整版.xlsx'
wb.save(output_file)

print(f"✅ Excel文件已生成: {output_file}")
print(f"📊 总产品数: {len(products)}")

# 统计品牌分布
brand_count = {}
for product in products:
    brand = product.get('品牌', 'Unknown')
    brand_count[brand] = brand_count.get(brand, 0) + 1

print("\n品牌分布:")
for brand, count in sorted(brand_count.items(), key=lambda x: x[1], reverse=True):
    print(f"  {brand}: {count}")

cursor.close()
conn.close()
